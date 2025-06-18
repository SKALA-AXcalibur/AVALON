import openpyxl
from fastapi import UploadFile
from io import BytesIO
from dto.request.spec.db import DbDesignDto, ColumnDto
from typing import Dict, List
import logging


class DbDesignParserService:
    """
    테이블 설계서(엑셀) 파일을 파싱하여 DB 설계 정보를 구조화하는 서비스 클래스
    """

    def find_custom_header_row(self, sheet, required_keywords):
        """
        엑셀 시트 내에서 '필수 키워드'가 모두 포함된 헤더 행을 탐색
        """
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            try:
                if all(
                    any(keyword in cell for cell in row_values)
                    for keyword in required_keywords
                ):
                    return idx, row
            except Exception as e:
                logging.warning(f"[헤더 행 추출 중 예외] row: {row_values}, error: {e}")
                continue
        return None, None

    def get_header_map(self, header_row):
        """
        한글 헤더명 → 내부 속성명 매핑용 딕셔너리 생성
        """
        header_map = {
            "테이블 한글명": "name",
            "컬럼 한글명": "col_name",
            "데이터타입": "type",
            "길이": "length",
            "PK": "is_pk",
            "FK": "fk",
            "Null여부": "is_null",
            "제약 조건": "const",
            "설명": "desc",
        }
        return {
            i: header_map.get(str(header_row[i]), str(header_row[i]))
            for i in range(len(header_row))
        }

    def get_merged_cell_value(self, sheet, row_idx, col_idx):
        """
        병합 셀 보정: 주어진 좌표가 병합 셀이라면 좌상단 값 반환, 아니면 원래 값 반환
        """
        try:
            cell = sheet.cell(row=row_idx, column=col_idx)
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    top_left_cell = sheet.cell(
                        row=merged_range.min_row, column=merged_range.min_col
                    )
                    return top_left_cell.value
            return cell.value
        except Exception as e:
            logging.warning(
                f"[병합 셀 값 추출 예외] row={row_idx}, col={col_idx}, error: {e}"
            )
            return None

    async def parse_dbdesign_file(self, upload_file: UploadFile) -> List[DbDesignDto]:
        """
        테이블 정의서(컬럼) 엑셀 파일 파싱 후, DB 설계 DTO 리스트 반환
        """
        try:
            contents = await upload_file.read()
            workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
            sheet = workbook["테이블정의서(컬럼)"]
        except Exception as e:
            logging.warning(f"[엑셀 파일 오픈 실패] error: {e}")
            return []

        required_keywords = ["Null여부", "데이터타입"]
        try:
            header_idx, header_row = self.find_custom_header_row(
                sheet, required_keywords
            )
        except Exception as e:
            logging.warning(f"[헤더 행 탐색 실패] error: {e}")
            return []
        if header_row is None:
            logging.warning("[헤더 미발견] 필수 헤더 행을 찾지 못했습니다.")
            return []

        # 병합 셀 보정
        real_header = []
        for i in range(1, sheet.max_column + 1):
            try:
                val = self.get_merged_cell_value(sheet, header_idx, i)
                real_header.append(val if val is not None else f"COL{i}")
            except Exception as e:
                logging.warning(f"[헤더 병합 셀 추출 실패] idx={i}, error: {e}")
                real_header.append(f"COL{i}")

        header_map = self.get_header_map(real_header)
        table_map: Dict[str, List[ColumnDto]] = {}

        # 본문 데이터 행 순회
        for row_idx in range(header_idx + 1, sheet.max_row + 1):
            try:
                row_dict = {}
                empty_row = True
                for col_idx in range(1, sheet.max_column + 1):
                    key = header_map.get(col_idx - 1)
                    if not key:
                        continue
                    value = self.get_merged_cell_value(sheet, row_idx, col_idx)
                    row_dict[key] = value
                    if value is not None:
                        empty_row = False
                if empty_row:
                    continue 

                # 테이블명 누락된 행은 skip
                raw_table_name = row_dict.get("name")
                if not raw_table_name or not str(raw_table_name).strip():
                    continue
                table_name = str(raw_table_name).strip()

                # 컬럼 데이터
                column_data = {
                    "col_name": str(row_dict.get("col_name") or ""),
                    "type": str(row_dict.get("type") or ""),
                    "length": (
                        int(row_dict.get("length")) if row_dict.get("length") else None
                    ),
                    "is_pk": str(row_dict.get("is_pk") or "").strip().upper() in ["Y", "PK"],
                    "fk": str(row_dict.get("fk")) if row_dict.get("fk") else None,
                    "is_null": {
                                "Y": True,
                                "N": False
                            }.get(str(row_dict.get("is_null") or "").strip().upper(), True),
                    "const": (
                        str(row_dict.get("const"))
                        if row_dict.get("const")
                        else None
                    ),
                    "desc": str(row_dict.get("desc")) if row_dict.get("desc") else None,
                }
                column_dto = ColumnDto(**column_data)
                # 테이블별로 컬럼 추가
                table_map.setdefault(table_name, []).append(column_dto)
            except Exception as e:
                logging.warning(f"[데이터 행 파싱 실패] row={row_idx}, error: {e}")
                continue  # 에러 발생 시 해당 행만 스킵, 전체 파싱은 계속

        db_designs = [
            DbDesignDto(name=tbl_name, column=columns)
            for tbl_name, columns in table_map.items()
        ]
        return db_designs
