# ai/service/spec/requirement_parser.py
import logging
import openpyxl
from fastapi import UploadFile
from io import BytesIO
from dto.request.spec.requirement import Requirement
from config import config as cfg


class RequirementParserService:
    """
    요구사항정의서(엑셀) 파일을 파싱하여 Requirement 객체 리스트로 반환하는 서비스
    """

    def find_custom_header_row(self, sheet, required_keywords):
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            row_values = [str(cell or "") for cell in row]
            if all(
                any(keyword in cell for cell in row_values)
                for keyword in required_keywords
            ):
                return idx, row
        return None, None

    def get_header_map(self, header_row):

        return {
            i: cfg.REQ_HEADER_MAP.get(str(header_row[i]), str(header_row[i]))
            for i in range(len(header_row))
        }

    async def parse_requirement_file(self, upload_file: UploadFile):
        """
        요구사항정의서 파싱, 객체 리스트 반환
        """
        try:
            # 파일 바이너리 읽어 openpyxl 워크북 로드
            contents = await upload_file.read()
            workbook = openpyxl.load_workbook(BytesIO(contents), data_only=True)
            sheet = workbook[cfg.REQ_SHEET_NAME]
        except Exception as e:
            logging.warning(f"[엑셀 파일 오픈/시트 접근 실패] error: {e}")
            return []

        # 헤더 행 자동 탐색 (대분류, 중분류, 소분류 포함)
        try:
            header_idx, header_row = self.find_custom_header_row(
                sheet, cfg.REQ_REQUIRED_KEYWORDS
            )
        except Exception as e:
            logging.warning(f"[헤더 탐색 중 예외] error: {e}")
            return []
        if header_row is None:
            logging.warning("[헤더 미발견] 필수 컬럼 포함된 헤더를 찾지 못함.")
            return []

        header_map = self.get_header_map(header_row)
        requirements = []
        for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            try:
                # 헤더 행 이전은 skip
                if idx <= header_idx:
                    continue
                # 완전히 빈 행은 skip
                if all(cell is None for cell in row):
                    continue
                # 각 열 값 → 내부 필드 dict 변환
                row_dict = {header_map[i]: row[i] for i in range(len(header_row))}
                requirement_data = {
                    field: row_dict.get(field)  # default None은 생략 가능
                    for field in cfg.REQUIRED_FIELDS
                }
                requirements.append(Requirement(**requirement_data))
            except Exception as e:
                logging.warning(f"[요구사항 행 파싱 실패] idx={idx}, error: {e}")
                continue  # 오류 발생 시 해당 행만 스킵

        return requirements
